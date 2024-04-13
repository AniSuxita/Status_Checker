
import time
import requests
import aiohttp
import asyncio
import openpyxl
from bs4 import BeautifulSoup

count = 0
sitemap_list = []
start = time.time()

async def fetch_all_webpage_urls(session, url):
    tasks = []
    res = requests.get(url)
    xml_soup = BeautifulSoup(res.content, 'xml')

    sitemap_atr = xml_soup.find_all('sitemap')
    if sitemap_atr:
        for item in sitemap_atr:
            loc_urls = item.find('loc').text
            sitemap_list.append(loc_urls)
    else:
        sitemap_list.append(url)

    for sitemap_url in sitemap_list[:2]:
        async with session.get(sitemap_url) as response:
            content = await response.text()
            webpage_xml_soup = BeautifulSoup(content, 'xml')
            for link in webpage_xml_soup.find_all('loc'):
                webpage_url = link.text
                tasks.append(webpage_url)
    return tasks


async def fetch_all_webpage_statuses(url, save_path):
    global count
    statuses = []
    async with aiohttp.ClientSession() as session:
        webpage_urls = await fetch_all_webpage_urls(session, url)
        statuses = await asyncio.gather(*[fetch_webpage_status(session, webpage_url) for webpage_url in webpage_urls])

        # statuses = [status for status in statuses if status is not None]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Webpage Statuses"

        ws['A1'] = 'URL'
        ws['B1'] = 'Status'

        for row, (url, status) in enumerate(zip(webpage_urls, statuses), start=2):
            ws[f'A{row}'] = url
            ws[f'B{row}'] = status

        wb.save(save_path)


async def fetch_webpage_status(session, webpage_url):
    global count
    try:
        async with session.get(webpage_url, timeout=10) as response:
            webpage_res = response.status
            print(webpage_url, webpage_res)
            count += 1
            return webpage_res
    except asyncio.TimeoutError:
        print(f"Timeout error for URL: {webpage_url}")
        return None
    except Exception as e:
        print(f"Error fetching URL {webpage_url}: {e}")
        return None


end = time.time()
total = end - start
print(total)